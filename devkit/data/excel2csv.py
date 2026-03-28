"""Excel to CSV converter."""

import csv
import os
from typing import List, Optional


def list_sheets(file: str) -> List[str]:
    """List sheet names in an Excel file."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required: pip install devkit-tools[data]")
    wb = load_workbook(file, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def excel_to_csv(file: str, output: str, sheet: Optional[str] = None) -> None:
    """Convert an Excel sheet to CSV."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required: pip install devkit-tools[data]")
    wb = load_workbook(file, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    os.makedirs(os.path.dirname(output) or ".", exist_ok=True)
    with open(output, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
    wb.close()
